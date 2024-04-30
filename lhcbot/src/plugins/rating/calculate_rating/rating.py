import os
import openpyxl
import re

contest_list = ["2024_90"]

# 获取学生信息列表
def get_stu_info():
    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\information.xlsx"
    workbook  = openpyxl.load_workbook(path) 
    sheet = workbook.active # 打开活动表单
    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1 # 获取学生数量
    stu = {} # 学生信息字典
    stu_list = {} # 学生信息字典集合
    for i in range(2, int(stu_num) + 2):
        stu.clear()
        name = sheet.cell(row=i, column=2).value
        stu['luogu_name'] = sheet.cell(row=i, column=3).value
        stu['at_id'] = sheet.cell(row=i, column=4).value
        stu['cf_id'] = sheet.cell(row=i, column=5).value
        stu_list[name] = stu.copy()
    return stu_list


# 获取学生总分数列表
def get_stu_all_rating():
    stu_info = get_stu_info()

    # 获取当前文件所在的文件夹路径
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\rating.xlsx"
    workbook  = openpyxl.load_workbook(path) 
    sheet = workbook.active # 打开活动表单
    student_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1 # 获取学生数量
    cur_stu = {} # 学生信息字典
    stu_dic = {} # 学生信息字典序列
    for i in range(2, int(student_num) + 2):
        cur_stu.clear()
        name = sheet.cell(row=i, column=1).value
        if name not in stu_info.keys(): # 如果学生信息表中没有该学生信息，则跳过
            continue
        cur_stu['rating'] = sheet.cell(row=i, column=2).value
        cur_stu['at_rating'] = sheet.cell(row=i, column=3).value
        cur_stu['cf_rating'] = sheet.cell(row=i, column=4).value
        stu_dic[name] = cur_stu.copy()
    workbook.save(path)
    return stu_dic


# 获取学生比赛分数变化 
def get_rating_change_info(x):
    current_file_path = os.path.abspath(__file__)
    current_folder_path = os.path.dirname(current_file_path)
    path = current_folder_path + "\\files\\rating_change\\" + str(x) + ".xlsx"
    try:
        workbook  = openpyxl.load_workbook(path)
    except:
        print("找不到名称为 " + str(x) + " 的比赛")
        return {}
    sheet = workbook.active
    stu_num = int(re.findall(r'\d+', sheet.dimensions)[-1]) - 1
    stu_dic = {}
    for i in range(2, int(stu_num) + 2):
        cur_stu = {}
        rank = sheet.cell(row=i, column=1).value
        name = sheet.cell(row=i, column=2).value
        score = sheet.cell(row=i, column=3).value
        old_rating = sheet.cell(row=i, column=4).value
        delta = sheet.cell(row=i, column=5).value
        new_rating = sheet.cell(row=i, column=6).value
        cur_stu['rank'] = rank
        cur_stu['name'] = name
        cur_stu['score'] = score
        cur_stu['old_rating'] = old_rating
        cur_stu['delta'] = delta
        cur_stu['new_rating'] = new_rating
        stu_dic[name] = cur_stu.copy()
    return stu_dic


def QQ_ask(s):
    s = s[1:]
    s = s.strip()
    print(s)

    # 单命令查询
    if s == "at" or s == "AT":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            cur_stu = {}
            cur_stu['name'] = x
            cur_stu['at_rating'] = stu_all_rating[x]['at_rating']
            stu_list.append(cur_stu.copy())
        stu_list.sort(key=lambda x: x['at_rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['at_rating']}\n"
        return res
    
    if s == "cf" or s == "CF":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            cur_stu = {}
            cur_stu['name'] = x
            cur_stu['cf_rating'] = stu_all_rating[x]['cf_rating']
            stu_list.append(cur_stu.copy())
        stu_list.sort(key=lambda x: x['cf_rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['cf_rating']}\n"
        return res
    
    if s == "rating" or s == "Rating":
        res = ""
        stu_all_rating = get_stu_all_rating()
        stu_list = []
        for x in stu_all_rating:
            cur_stu = {}
            cur_stu['name'] = x
            cur_stu['rating'] = stu_all_rating[x]['rating']
            stu_list.append(cur_stu.copy()) 
        stu_list.sort(key=lambda x: x['rating'], reverse=True)
        for x in stu_list:
            res += f"{x['name']} : {x['rating']}\n"
        return res
    
    # 多命令查询
    list_s= s.split(" ")
    if list_s[0] == "contest":
        if list_s[1] in contest_list:
            res = ""
            stu_rating_change = get_rating_change_info(list_s[1])
            stu_list = []
            for x in stu_rating_change:
                cur_stu = {}
                cur_stu['name'] = x
                cur_stu['rank'] = stu_rating_change[x]['rank']
                cur_stu['delta'] = stu_rating_change[x]['delta']
                stu_list.append(cur_stu.copy())
            stu_list.sort(key=lambda x: x['rank'])
            for x in stu_list:
                res += f"{x['rank']} : {x['name']}   {x['delta']}\n"
            return res
        else:
            return "没有找到该比赛"

    return "error"
